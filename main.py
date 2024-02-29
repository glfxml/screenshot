#!/usr/env python3
# -*- coding: UTF-8 -*-
import os
import time
import pyautogui  # 右键菜单元素选择
from PIL import Image
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains

# 1. 读取 picid.xlsx 文件, 获取作品id
# 2. 拼接作品id 成 url
# 3. 访问 url, 并把页面内容转成英文截图, 保存到 data 目录下的 pic 目录下

# 读取 picid.xlsx 文件, 获取作品id
def get_picid(file):
    # 读取 xlsx 文件
    wb = load_workbook(file)
    # 选择工作表
    sheet = wb['Sheet1']

    list = []
    # 遍历工作表的每一行
    for row in sheet.iter_rows(values_only=True, min_row=2):
        picid = row[0]
        list.append(picid)

    return list

# selenium 模拟浏览器访问 url, 并把页面内容转成英文截图, 保存到 data 目录下的 pic 目录下
def scroll_and_screenshot(url, id):
    chrome_options = Options()
    chrome_options.add_argument('--lang=en')
    # 启动WebDriver并打开页面
    driver = webdriver.Chrome(chrome_options)
    driver.get(url)
    # 设置浏览器窗口大小
    driver.maximize_window()
    # 隐式等待，暂时可以先不用管
    driver.implicitly_wait(10)
    # 可以设置超时时间避免需要等待很长时间，但是可能会导致元素还没加载出来，可以根据自己的网络对超时时间进行调整
    driver.set_page_load_timeout(6)

    # 等待页面加载
    time.sleep(3)
    # 关闭登录弹窗
    driver.find_element(by=By.CLASS_NAME, value='jsCtr_login_pop_close').click()

    # 通过JS语句设置元素属性style arguments是参数
    path = driver.find_element(by=By.CLASS_NAME, value='unlogin-block')
    js_style = "arguments[0].setAttribute('style', arguments[1]);"
    driver.execute_script(js_style, path, "display: none;")

    # 通过JS语句设置元素属性style arguments是参数
    detailHead = driver.find_element(by=By.CLASS_NAME, value='wt-detailHead')
    jsStyle = "arguments[0].setAttribute('style', arguments[1]);"
    driver.execute_script(jsStyle, detailHead, "display: none;")

    # 隐藏以图搜图弹窗
    detailHead = driver.find_element(by=By.CLASS_NAME, value='wt-blueTipsWrap')
    jsStyle = "arguments[0].setAttribute('style', arguments[1]);"
    driver.execute_script(jsStyle, detailHead, "display: none;")

    # 通过JS语句设置元素属性style arguments是参数
    js = "arguments[0].setAttribute('style', arguments[1]);"
    # css语句，给元素添加边框
    style = "border: 5px solid red;"
    elements = driver.find_elements(By.PARTIAL_LINK_TEXT, 'i')
    for index in range(len(elements)):
        # 执行JS语句，将元素作为参数传递
        driver.execute_script(js, elements[index], style)

    # 右键点击页面
    body = driver.find_element("xpath", "//div[@class='wt-detailsContainer']")
    action = ActionChains(driver)
    action.context_click().perform()
    rightClick = ActionChains(driver)  # 实例化ActionChains类
    rightClick.context_click(body).perform()  # context_click(body)在body上执行右键操作，perform()是一个执行动作
    pyautogui.typewrite(['down'] * 8)  # 移动到菜单中的翻译选项
    pyautogui.typewrite(["enter"])  # 选择翻译选项

    time.sleep(5)

    # 获取网页高度
    body_height = driver.execute_script('return document.body.scrollHeight;')
    window_height = driver.execute_script('return window.innerHeight;')
    js = "window.scroll(0,arguments[0]*arguments[1])"

    img_path = f"J:\\screenshot\\data\\pic\\{id}"
    os.makedirs(img_path, exist_ok=True)

    i = 0
    image_paths = []
    while i * window_height < body_height:
        driver.execute_script(js, window_height, i)
        out_file = f"J:\\screenshot\\data\\pic\\{id}\\{id}_{i}.png"
        image_paths.append(out_file)
        driver.get_screenshot_as_file(out_file)
        i += 1
        if i > 1:
            break

    driver.quit()

    # 拼接图片
    out_img = img_path + ".png"
    concatenate_images(image_paths, out_img)
    clear_images(image_paths)
    # 删除目录
    os.rmdir(img_path)


def clear_images(image_paths):
    for image_path in image_paths:
        os.remove(image_path)


# 拼接图片
def concatenate_images(image_paths, output_path):
    images = [Image.open(image_path) for image_path in image_paths]
    # 获取图片列表中第一张图片的尺寸
    width, height = images[0].size

    # 创建一个新的图像对象，宽度不变，高度为所有图像的高度之和
    result = Image.new("RGB", (width, sum(img.height for img in images)))

    # 将每张图片依次粘贴到新图像上
    y_offset = 0
    for img in images:
        result.paste(img, (0, y_offset))
        y_offset += img.height

    # 保存拼接后的图像
    result.save(output_path)

    # 关闭所有图像对象
    for img in images:
        img.close()


def main():
    picid = get_picid(r"J:\screenshot\data\xlsx\picid.xlsx")
    for id in picid:
        url = "https://weili.ooopic.com/weili_" + str(id) + ".html"
        scroll_and_screenshot(url, id)


if __name__ == '__main__':
    main()
